# -*- coding: utf-8 -*-
"""
Created on Fri May 15 10:33:20 2020
Title: 
@author: Dr. Tian Guo
"""
import openpyxl
from openpyxl import load_workbook
import numpy as np 
import xlrd 
import torch 

def load_data(flag, file="/home/zq/program/ORR_prediction-master/DFT_data/my_data_2.xlsx"):
    # wb = xlrd.open_workbook(file)
    # sheet = wb.sheet_by_index(0)

    workbook_object = load_workbook(file)
    names = workbook_object.sheetnames
    sheet = workbook_object.worksheets[0]
    
    num = 29

    d_orbital_of_metal = np.array([sheet.cell(loopi + 2, 2).value for loopi in range(num)])         # 5
    group = np.array([sheet.cell(loopi + 2, 3).value for loopi in range(num)])                      # 7
    electronegativity = np.array([sheet.cell(loopi+2,4).value for loopi in range(num)])             # 1

    nearrest_N = np.array([sheet.cell(loopi+2,5).value for loopi in range(num)])                     # 8
    nearrest_C = np.array([sheet.cell(loopi+2,6).value for loopi in range(num)])                    # 10

    V_ORR_E_O = np.array([sheet.cell(loopi+2,7).value for loopi in range(19)])                      # 11
    V_ORR_E_H = np.array([sheet.cell(loopi+2,8).value for loopi in range(19)])                      # 12


    X_data = np.stack((d_orbital_of_metal, group, electronegativity,
                       nearrest_N, nearrest_C), axis=0).T
    X_data = torch.from_numpy(X_data).float()  
    
    if flag == "E_O":
        y_data = torch.from_numpy(V_ORR_E_O).float()
    else: 
        y_data = torch.from_numpy(V_ORR_E_H).float()
    
    return X_data, y_data, [d_orbital_of_metal, group, electronegativity,
                            nearrest_N, nearrest_C, y_data]

    
def build_dataset(flag):
    X_data, y_data, _ = load_data(flag)
    
    train_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]
    # if flag == "1V":
    test_list = [6, 12, 18]
    # else: 
    #     test_list = [5, 11]  
        
    for idx in test_list: 
        train_list.remove(idx)
    
    predict_list = [19+loopi for loopi in range(10)]
    
    X_data_train, y_data_train = [], []
    X_data_test, y_data_test = [], [] 
    X_data_predict = [] 
    for index in range(29):  
        if index in train_list:
            X_data_train.append(X_data[index,:].reshape(1,1,5).float())
            y_data_train.append(y_data[index].reshape(1)) 
            
        if index in test_list: 
            X_data_test.append(X_data[index,:].reshape(1,1,5).float())
            y_data_test.append(y_data[index].reshape(1)) 
            
        if index in predict_list:
            X_data_predict.append(X_data[index,:].reshape(1,1,5).float()) 
            
    return X_data_train, y_data_train, X_data_test, y_data_test, X_data_predict


if __name__ == "__main__": 
    flag = "E_O"
    X_data_train, y_data_train, X_data_test, y_data_test, X_data_predict = build_dataset(flag) 
    
